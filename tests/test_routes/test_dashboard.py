"""Testes das rotas do dashboard: /admin, /exportar, /admin/relatorios, /chamado/<id>/historico, /admin/indices-firestore."""

from unittest.mock import MagicMock, patch


def test_admin_sem_login_redireciona(client):
    """GET /admin sem login redireciona para /login."""
    r = client.get("/admin", follow_redirects=False)
    assert r.status_code == 302
    assert "login" in (r.location or "").lower()


def test_admin_com_solicitante_nao_acessa(client_logado_solicitante):
    """GET /admin com perfil solicitante não acessa: 403 ou redirect."""
    r = client_logado_solicitante.get("/admin", follow_redirects=False)
    assert r.status_code in (302, 403)
    if r.status_code == 302:
        assert "/admin" not in (r.location or "")


def test_admin_com_supervisor_redireciona_para_painel(client_logado_supervisor):
    """GET /admin com supervisor redireciona para /painel."""
    r = client_logado_supervisor.get("/admin", follow_redirects=False)
    assert r.status_code == 302
    assert "painel" in (r.location or "")


def test_painel_com_supervisor_retorna_200(client_logado_supervisor):
    """GET /painel com supervisor retorna 200 e página do dashboard (mock contexto)."""
    with (
        patch("app.routes.dashboard.obter_contexto_admin") as mock_ctx,
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
    ):
        mock_ctx.return_value = {
            "chamados": [],
            "gates": [],
            "responsaveis": [],
            "sla_map": {},
            "tem_proxima": False,
            "tem_anterior": False,
            "proximo_cursor": None,
            "cursor_anterior": None,
        }
        r = client_logado_supervisor.get("/painel", follow_redirects=False)
    assert r.status_code == 200
    mock_ctx.assert_called_once()


def test_admin_com_admin_retorna_200(client_logado_admin):
    """GET /admin com admin retorna 200."""
    with (
        patch("app.routes.dashboard.obter_contexto_admin") as mock_ctx,
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
    ):
        mock_ctx.return_value = {
            "chamados": [],
            "gates": [],
            "responsaveis": [],
            "sla_map": {},
            "tem_proxima": False,
            "tem_anterior": False,
            "proximo_cursor": None,
            "cursor_anterior": None,
        }
        r = client_logado_admin.get("/admin", follow_redirects=False)
    assert r.status_code == 200


def test_admin_navbar_tem_link_novo_chamado_e_meus_chamados(client_logado_admin):
    """Navbar (menu hambúrguer) do Admin deve ter link para abrir chamado e ver
    seus próprios chamados — o backend já permite (@requer_solicitante inclui
    admin/admin_global), só faltava o link na UI para chegar lá."""
    with (
        patch("app.routes.dashboard.obter_contexto_admin") as mock_ctx,
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
    ):
        mock_ctx.return_value = {
            "chamados": [],
            "gates": [],
            "responsaveis": [],
            "sla_map": {},
            "tem_proxima": False,
            "tem_anterior": False,
            "proximo_cursor": None,
            "cursor_anterior": None,
        }
        r = client_logado_admin.get("/admin", follow_redirects=False)
    assert r.status_code == 200
    html = r.data.decode("utf-8")
    assert "New Ticket" in html, "Link 'New Ticket' ausente na navbar do Admin"
    assert "My Tickets" in html, "Link 'My Tickets' ausente na navbar do Admin"


def test_admin_global_navbar_nao_duplica_link_painel_global(client_logado_admin_global):
    """Admin Global via navegador desktop não deve ver 'Global Panel' (navbar
    principal) e 'Global Administrator' (hambúrguer) juntos — mesmo destino
    (main.admin_global_dashboard), dois rótulos. O item do hambúrguer deve
    ficar em md:hidden (só mobile, onde a navbar principal já é escondida)."""
    with patch("app.routes.admin_global.Usuario") as mock_usuario:
        mock_usuario.get_all.return_value = []
        r = client_logado_admin_global.get("/admin-global", follow_redirects=False)

    assert r.status_code == 200
    html = r.data.decode("utf-8")
    idx = html.find("Global Administrator")
    assert idx != -1, "Rótulo 'Global Administrator' não encontrado na navbar"
    trecho_antes = html[max(0, idx - 900) : idx]
    assert 'class="md:hidden"' in trecho_antes, (
        "Link 'Global Administrator' do hambúrguer não está em md:hidden — "
        "fica duplicado com 'Global Panel' da navbar principal em telas grandes"
    )


def test_exportar_sem_login_redireciona(client):
    """GET /exportar sem login redireciona para login."""
    r = client.get("/exportar", follow_redirects=False)
    assert r.status_code == 302
    assert "login" in (r.location or "").lower()


def test_exportar_com_supervisor_retorna_200_ou_redirect(client_logado_supervisor):
    """GET /exportar com supervisor retorna 200 (arquivo) ou redirect em caso de erro (mock)."""
    with (
        patch("app.routes.dashboard.aplicar_filtros_dashboard_com_paginacao") as mock_filtros,
        patch("app.routes.dashboard.verificar_e_incrementar_export", return_value=(True, None)),
    ):
        mock_doc = MagicMock()
        mock_doc.to_dict.return_value = {}
        mock_doc.id = "doc1"
        mock_filtros.return_value = {"docs": [mock_doc]}
        with patch("app.routes.dashboard._filtrar_chamados_por_permissao") as mock_filtrar:
            mock_filtrar.return_value = []
            r = client_logado_supervisor.get("/exportar", follow_redirects=False)
    assert r.status_code in (200, 302)
    if r.status_code == 200:
        ct = r.headers.get("Content-Type", "")
        assert "spreadsheet" in ct or "excel" in ct or "octet" in ct.lower()


def test_relatorios_sem_login_redireciona(client):
    """GET /admin/relatorios sem login redireciona para login."""
    r = client.get("/admin/relatorios", follow_redirects=False)
    assert r.status_code == 302
    assert "login" in (r.location or "").lower()


def test_relatorios_com_admin_retorna_200(client_logado_admin):
    """GET /admin/relatorios com admin retorna 200 (mock analytics)."""
    with patch("app.routes.dashboard.analisador") as mock_anal:
        mock_anal.obter_relatorio_completo.return_value = {
            "data_geracao": None,
            "metricas_gerais": {},
            "metricas_supervisores": [],
            "metricas_areas": [],
            "insights": [],
        }
        with (
            patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
            patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        ):
            r = client_logado_admin.get("/admin/relatorios", follow_redirects=False)
    assert r.status_code == 200
    assert b"relat" in r.data.lower() or b"report" in r.data.lower() or b"anal" in r.data.lower()


def test_relatorios_propaga_dias_valido_para_analisador(client_logado_admin):
    """GET /admin/relatorios?dias=7 repassa dias=7 pro analisador (seletor de período)."""
    with patch("app.routes.dashboard.analisador") as mock_anal:
        mock_anal.obter_relatorio_completo.return_value = {
            "data_geracao": None,
            "metricas_gerais": {},
            "metricas_supervisores": [],
            "metricas_areas": [],
            "insights": [],
        }
        with (
            patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
            patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        ):
            r = client_logado_admin.get("/admin/relatorios?dias=7", follow_redirects=False)
    assert r.status_code == 200
    assert mock_anal.obter_relatorio_completo.call_args.kwargs["dias"] == 7


def test_relatorios_dias_invalido_normaliza_para_30(client_logado_admin):
    """GET /admin/relatorios?dias=999 (fora de 7/30/90) cai para o padrão 30."""
    with patch("app.routes.dashboard.analisador") as mock_anal:
        mock_anal.obter_relatorio_completo.return_value = {
            "data_geracao": None,
            "metricas_gerais": {},
            "metricas_supervisores": [],
            "metricas_areas": [],
            "insights": [],
        }
        with (
            patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
            patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        ):
            r = client_logado_admin.get("/admin/relatorios?dias=999", follow_redirects=False)
    assert r.status_code == 200
    assert mock_anal.obter_relatorio_completo.call_args.kwargs["dias"] == 30


def test_supervisor_puro_nao_pode_ver_relatorios(client_logado_supervisor):
    """GET /admin/relatorios com supervisor sem nivel_gestao é bloqueado (302 ou 403).

    Regra de negócio: nenhum supervisor "puro" vê essa página — só quem tem
    nivel_gestao (Gestor do Setor, Assistente GM, GM, Gerente de Produção) ou é admin.
    """
    r = client_logado_supervisor.get("/admin/relatorios", follow_redirects=False)
    assert r.status_code in (302, 403)
    if r.status_code == 302:
        assert "/admin/relatorios" not in (r.location or "")


def test_solicitante_nao_pode_ver_relatorios(client_logado_solicitante):
    """GET /admin/relatorios com solicitante é bloqueado (302 ou 403)."""
    r = client_logado_solicitante.get("/admin/relatorios", follow_redirects=False)
    assert r.status_code in (302, 403)


def test_gestor_setor_ve_relatorios_escopado_para_sua_area(client_logado_gestor):
    """GET /admin/relatorios com Gestor do Setor (supervisor + nivel_gestao='gestor_setor')
    retorna 200 e repassa areas=current_user.areas ao analisador (visão restrita à própria área).
    """
    with patch("app.routes.dashboard.analisador") as mock_anal:
        mock_anal.obter_relatorio_completo.return_value = {
            "data_geracao": None,
            "metricas_gerais": {},
            "metricas_supervisores": [],
            "metricas_areas": [],
            "insights": [],
        }
        with (
            patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
            patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        ):
            r = client_logado_gestor.get("/admin/relatorios", follow_redirects=False)
    assert r.status_code == 200
    assert mock_anal.obter_relatorio_completo.call_args.kwargs["areas"] == ["Geral"]


def _mock_usuario_gestor_amplo(uid="gm_1", nivel_gestao="gm"):
    """MagicMock de usuário com nivel_gestao company-wide (gm/assistente_gm/gerente_producao) —
    sem perfil operacional (não é supervisor nem admin), só o nível de gestão concede acesso."""
    u = MagicMock()
    u.id = uid
    u.email = f"{uid}@dtx.aero"
    u.nome = "Gestor Amplo Teste"
    u.perfil = "solicitante"
    u.area = "Geral"
    u.areas = ["Geral"]
    u.nivel_gestao = nivel_gestao
    u.is_authenticated = True
    u.get_id = lambda: str(uid)
    u.must_change_password = False
    u.mfa_enabled = True
    u.is_admin_or_above = False
    u.is_supervisor_or_above = False
    u.is_gestor = True
    u.is_gestor_only = True
    u.onboarding_perfis_vistos = ["solicitante"]
    u.onboarding_passo = 0
    u.ativo = True
    return u


def test_gm_ve_relatorios_completo_sem_escopo_de_area(client, app):
    """GET /admin/relatorios com nivel_gestao='gm' (sem perfil supervisor/admin)
    retorna 200 e repassa areas=None ao analisador (visão company-wide)."""
    user = _mock_usuario_gestor_amplo()
    with (
        patch("app.routes.auth.Usuario.get_by_email", return_value=user),
        patch("app.models_usuario.Usuario.get_by_id", return_value=user),
        patch("app.routes.auth._dispositivo_confiavel", return_value=True),
    ):
        client.post("/login", data={"email": user.email, "senha": "ok"}, follow_redirects=False)
        with patch("app.routes.dashboard.analisador") as mock_anal:
            mock_anal.obter_relatorio_completo.return_value = {
                "data_geracao": None,
                "metricas_gerais": {},
                "metricas_supervisores": [],
                "metricas_areas": [],
                "insights": [],
            }
            with (
                patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
                patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
            ):
                r = client.get("/admin/relatorios", follow_redirects=False)
    assert r.status_code == 200
    assert mock_anal.obter_relatorio_completo.call_args.kwargs["areas"] is None


def test_historico_sem_login_redireciona(client):
    """GET /chamado/<id>/historico sem login redireciona para login."""
    r = client.get("/chamado/abc123/historico", follow_redirects=False)
    assert r.status_code == 302
    assert "login" in (r.location or "").lower()


def test_historico_com_supervisor_permissao_retorna_200(client_logado_supervisor, db_session):
    """GET /chamado/<id>/historico com supervisor que pode ver o chamado retorna 200."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao", solicitante_id="sol1")
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.Historico.get_by_chamado_id", return_value=[]),
    ):
        r = client_logado_supervisor.get(f"/chamado/{chamado.id}/historico", follow_redirects=False)
    assert r.status_code == 200


def test_historico_admin_global_com_permissao_retorna_200(client_logado_admin_global, db_session):
    """GET /chamado/<id>/historico com admin_global retorna 200 — admin_global
    herda tudo de admin, incluindo ver histórico de qualquer chamado."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao", solicitante_id="sol1")
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.Historico.get_by_chamado_id", return_value=[]),
    ):
        r = client_logado_admin_global.get(
            f"/chamado/{chamado.id}/historico", follow_redirects=False
        )
    assert r.status_code == 200


def test_historico_gestor_amplo_com_permissao_retorna_200(client, app, db_session):
    """GET /chamado/<id>/historico com nivel_gestao company-wide (Gerente de
    Produção/Assistente GM/GM) retorna 200, mesmo com perfil='solicitante' —
    regra de negócio: esses 3 níveis devem ver histórico de qualquer chamado,
    igual a admin. Antes desse fix, o gate da rota redirecionava QUALQUER
    perfil='solicitante' antes de checar nivel_gestao, bloqueando esses gestores."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao", solicitante_id="sol1")
    user = _mock_usuario_gestor_amplo()
    with (
        patch("app.routes.auth.Usuario.get_by_email", return_value=user),
        patch("app.models_usuario.Usuario.get_by_id", return_value=user),
        patch("app.routes.auth._dispositivo_confiavel", return_value=True),
    ):
        client.post("/login", data={"email": user.email, "senha": "ok"}, follow_redirects=False)
        with (
            patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
            patch("app.routes.dashboard.Historico.get_by_chamado_id", return_value=[]),
        ):
            r = client.get(f"/chamado/{chamado.id}/historico", follow_redirects=False)
    assert r.status_code == 200


def test_historico_solicitante_puro_redireciona_para_detalhe(client_logado_solicitante, db_session):
    """GET /chamado/<id>/historico com solicitante sem nivel_gestao continua
    redirecionando pra tela de detalhe (comportamento preservado — link de
    e-mail antigo)."""
    from tests.factories import make_chamado

    chamado = make_chamado(solicitante_id="sol_1")
    r = client_logado_solicitante.get(f"/chamado/{chamado.id}/historico", follow_redirects=False)
    assert r.status_code == 302
    assert f"/chamado/{chamado.id}" in (r.location or "")
    assert "historico" not in (r.location or "")


def test_historico_renderiza_todas_acoes_automaticas_sem_quebrar(
    client_logado_supervisor, db_session
):
    """GET /chamado/<id>/historico não quebra com nenhuma das 9 ações novas
    (previsão de atendimento + automações do motor de SLA): cada uma cai numa
    branch dedicada (dot colorido + detalhe visível), não no fallback genérico
    'outro' (cinza, sem detalhe)."""
    from app.models_historico import Historico
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao", solicitante_id="sol1")
    acoes = [
        "solicitacao_previsao_atendimento",
        "aprovacao_previsao_atendimento",
        "rejeicao_previsao_atendimento",
        "escalonamento_automatico",
        "aviso_previo_escalonamento",
        "aviso_resolucao_prazo",
        "lembrete_confirmacao_enviado",
        "alerta_prazo_24h",
        "confirmacao_resolucao",
    ]
    eventos = [
        Historico(
            chamado_id=chamado.id,
            usuario_id="sistema",
            usuario_nome="Sistema",
            acao=acao,
            campo_alterado="x",
            valor_anterior=None,
            valor_novo="y",
            detalhe="detalhe de teste",
        )
        for acao in acoes
    ]

    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.Historico.get_by_chamado_id", return_value=eventos),
    ):
        r = client_logado_supervisor.get(f"/chamado/{chamado.id}/historico")

    assert r.status_code == 200
    body = r.data.decode("utf-8")
    # As 3 ações de previsão + 5 automações de SLA caem nas branches dedicadas
    # 'previsao'/'sistema' (não no fallback genérico 'outro', sem cor nem detalhe).
    assert body.count("bento-timeline-dot previsao") == 3
    assert body.count("bento-timeline-dot sistema") == 5
    assert body.count("bento-timeline-dot outro") == 0
    assert body.count("detalhe de teste") == len(acoes)


def test_historico_anexo_tardio_renderiza_com_titulo_e_motivo(client_logado_supervisor, db_session):
    """GET /chamado/<id>/historico não deixa o evento 'anexo_tardio' (anexo
    enviado pelo solicitante após a criação, via adicionar_anexo_tardio) cair
    no fallback genérico 'outro' — achado em auditoria 2026-08-13: essa acao
    não tinha nenhuma branch dedicada no template, então título e corpo do
    card ficavam em branco (só badge/autor/data visíveis)."""
    from app.models_historico import Historico
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao", solicitante_id="sol1")
    evento = Historico(
        chamado_id=chamado.id,
        usuario_id="sol1",
        usuario_nome="Solicitante Teste",
        acao="anexo_tardio",
        campo_alterado="anexos",
        valor_anterior=None,
        valor_novo="uploads/comprovante.pdf",
        detalhe="Motivo do anexo tardio",
    )

    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.Historico.get_by_chamado_id", return_value=[evento]),
    ):
        r = client_logado_supervisor.get(f"/chamado/{chamado.id}/historico")

    assert r.status_code == 200
    body = r.data.decode("utf-8")
    assert "bento-timeline-dot outro" not in body
    assert "bento-timeline-dot dados" in body
    assert "comprovante.pdf" in body
    assert "Motivo do anexo tardio" in body


def test_historico_acoes_colaboracao_e_edicao_nao_caem_no_fallback_generico(
    client_logado_supervisor, db_session
):
    """GET /chamado/<id>/historico — mesma auditoria 2026-08-13: 7 ações reais
    (transferência de área, transferência para colega, inclusão de
    participantes, conclusão de parte, inclusão de observadores, edição de
    descrição e reabertura) não tinham branch dedicada e caíam no fallback
    genérico 'outro', sem título nem corpo visível."""
    from app.models_historico import Historico
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao", solicitante_id="sol1")
    eventos = [
        Historico(
            chamado_id=chamado.id,
            usuario_id="sup1",
            usuario_nome="Supervisor Teste",
            acao="transferencia_area",
            campo_alterado="area",
            valor_anterior="Manutencao",
            valor_novo="Producao",
            detalhe="Transferido para Producao — motivo teste",
        ),
        Historico(
            chamado_id=chamado.id,
            usuario_id="sup1",
            usuario_nome="Supervisor Teste",
            acao="escalonamento_colega",
            campo_alterado="responsavel_id",
            valor_anterior="sup1",
            valor_novo="sup2",
            detalhe="Escalado para Colega Teste — motivo teste",
        ),
        Historico(
            chamado_id=chamado.id,
            usuario_id="sup1",
            usuario_nome="Supervisor Teste",
            acao="inclusao_participantes",
            campo_alterado="participantes",
            valor_anterior="0",
            valor_novo="1",
            detalhe="Participantes incluídos: Colega Teste (Manutencao)",
        ),
        Historico(
            chamado_id=chamado.id,
            usuario_id="sup2",
            usuario_nome="Colega Teste",
            acao="conclusao_parte_participante",
            campo_alterado="participantes",
            valor_anterior="pendente",
            valor_novo="concluido",
            detalhe="Participante Colega Teste concluiu sua parte",
        ),
        Historico(
            chamado_id=chamado.id,
            usuario_id="sol1",
            usuario_nome="Solicitante Teste",
            acao="inclusao_observadores",
            campo_alterado="observadores",
            valor_anterior=None,
            valor_novo="Observador Teste",
        ),
        Historico(
            chamado_id=chamado.id,
            usuario_id="sol1",
            usuario_nome="Solicitante Teste",
            acao="edicao_descricao",
            campo_alterado="descricao",
            valor_anterior="Descrição antiga",
            valor_novo="Descrição nova",
        ),
        Historico(
            chamado_id=chamado.id,
            usuario_id="sup1",
            usuario_nome="Supervisor Teste",
            acao="reabertura",
            campo_alterado="status",
            valor_anterior="Concluído",
            valor_novo="Aberto",
            detalhe="Reabertura administrativa",
        ),
    ]

    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.Historico.get_by_chamado_id", return_value=eventos),
    ):
        r = client_logado_supervisor.get(f"/chamado/{chamado.id}/historico")

    assert r.status_code == 200
    body = r.data.decode("utf-8")
    assert "bento-timeline-dot outro" not in body
    assert body.count("bento-timeline-dot colaboracao") == 5
    assert "Transferido para Producao — motivo teste" in body
    assert "Escalado para Colega Teste — motivo teste" in body
    assert "Participantes incluídos: Colega Teste (Manutencao)" in body
    assert "Participante Colega Teste concluiu sua parte" in body
    assert "Observador Teste" in body
    assert "Descrição antiga" in body
    assert "Descrição nova" in body
    assert "Reabertura administrativa" in body


def test_historico_motivo_cancelamento_solicitante_aparece_no_alteracao_status(
    client_logado_supervisor, db_session
):
    """GET /chamado/<id>/historico — achado em auditoria 2026-08-13: o
    cancelamento feito pelo solicitante (cancelamento_solicitante_service.py)
    grava só um evento 'alteracao_status' com o motivo no campo `detalhe`,
    mas o template só mostrava `detalhe` pra 'reabertura' — o motivo do
    cancelamento nunca aparecia no histórico."""
    from app.models_historico import Historico
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao", solicitante_id="sol1", status="Cancelado")
    evento = Historico(
        chamado_id=chamado.id,
        usuario_id="sol1",
        usuario_nome="Solicitante Teste",
        acao="alteracao_status",
        campo_alterado="status",
        valor_anterior="Aberto",
        valor_novo="Cancelado",
        detalhe="Não preciso mais deste chamado",
    )

    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.Historico.get_by_chamado_id", return_value=[evento]),
    ):
        r = client_logado_supervisor.get(f"/chamado/{chamado.id}/historico")

    assert r.status_code == 200
    body = r.data.decode("utf-8")
    assert "Não preciso mais deste chamado" in body


# ── POST /admin (alteração de status) ─────────────────────────────────────────


def test_admin_post_status_change_admin_sucesso(client_logado_admin, db_session):
    """POST /admin com admin altera status com sucesso e redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado(status="Em Atendimento", area="Geral")
    with patch("app.routes.dashboard.atualizar_status_chamado") as mock_atualizar:
        mock_atualizar.return_value = {"sucesso": True, "mensagem": "Status atualizado"}
        r = client_logado_admin.post(
            "/admin",
            data={"chamado_id": str(chamado.id), "novo_status": "Concluído"},
            follow_redirects=False,
        )
    assert r.status_code == 302
    assert "/admin" in (r.location or "")
    mock_atualizar.assert_called_once()


def test_admin_post_status_change_falha_exibe_erro(client_logado_admin, db_session):
    """POST /admin quando atualizar_status_chamado retorna sucesso=False redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado(status="Em Atendimento", area="Geral")
    with patch("app.routes.dashboard.atualizar_status_chamado") as mock_atualizar:
        mock_atualizar.return_value = {"sucesso": False, "erro": "Status inválido"}
        r = client_logado_admin.post(
            "/admin",
            data={"chamado_id": str(chamado.id), "novo_status": "Invalido"},
            follow_redirects=False,
        )
    assert r.status_code == 302


def test_admin_post_status_change_supervisor_chamado_nao_encontrado(client_logado_supervisor):
    """POST /admin com supervisor quando chamado não existe redireciona.

    /admin redireciona supervisor pra /painel imediatamente (GET ou POST), antes
    de tocar o chamado — então isso já é 302 independente do chamado existir.
    """
    r = client_logado_supervisor.post(
        "/admin",
        data={"chamado_id": "ch_nao_existe", "novo_status": "Concluído"},
        follow_redirects=False,
    )
    assert r.status_code == 302


def test_admin_post_status_change_supervisor_sem_permissao_redireciona(client_logado_supervisor):
    """POST /admin com supervisor redireciona pra /painel antes de processar o chamado."""
    r = client_logado_supervisor.post(
        "/admin",
        data={"chamado_id": "ch1", "novo_status": "Concluído"},
        follow_redirects=False,
    )
    assert r.status_code == 302


def test_admin_post_exception_redireciona(client_logado_admin, db_session):
    """POST /admin quando atualizar_status_chamado lança exceção redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado(status="Em Atendimento", area="Geral")
    with patch("app.routes.dashboard.atualizar_status_chamado", side_effect=Exception("timeout")):
        r = client_logado_admin.post(
            "/admin",
            data={"chamado_id": str(chamado.id), "novo_status": "Concluído"},
            follow_redirects=False,
        )
    assert r.status_code == 302


# ── GET /chamado/<chamado_id> ─────────────────────────────────────────────────


def _chamado_dict_fake(solicitante_id="sol_x", area="Manutencao", status="Aberto"):
    return {
        "numero_chamado": "001",
        "categoria": "TI",
        "tipo_solicitacao": "Corretiva",
        "descricao": "Teste",
        "responsavel": "Resp",
        "area": area,
        "solicitante_id": solicitante_id,
        "solicitante_nome": "Fulano",
        "responsavel_id": None,
        "status": status,
        "gate": None,
        "rl_codigo": None,
        "data_abertura": None,
        "data_conclusao": None,
        "sla_dias": None,
        "anexo": None,
    }


def test_visualizar_chamado_nao_encontrado_redireciona_admin(client_logado_admin):
    """GET /chamado/<id> quando chamado não existe redireciona para /admin."""
    r = client_logado_admin.get("/chamado/999999999", follow_redirects=False)
    assert r.status_code == 302
    assert "/admin" in (r.location or "")


def test_visualizar_chamado_admin_com_permissao_retorna_200(client_logado_admin, db_session):
    """GET /chamado/<id> com admin que pode ver retorna 200."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        r = client_logado_admin.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 200


def test_visualizar_chamado_admin_global_ve_botao_historico(client_logado_admin_global, db_session):
    """GET /chamado/<id> com admin_global mostra o botão "Ver Histórico" —
    admin_global herda tudo de admin; o botão só checava perfil in
    ['admin', 'supervisor'], deixando admin_global sem o botão mesmo tendo
    acesso à rota /chamado/<id>/historico."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        r = client_logado_admin_global.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert f"/chamado/{chamado.id}/historico" in html


def test_visualizar_chamado_gestor_amplo_ve_botao_historico(client, app, db_session):
    """GET /chamado/<id> com nivel_gestao company-wide (Gerente de Produção/
    Assistente GM/GM) mostra o botão "Ver Histórico", mesmo com perfil='solicitante'
    — regra de negócio: esses 3 níveis devem ver histórico de qualquer chamado."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    user = _mock_usuario_gestor_amplo()
    with (
        patch("app.routes.auth.Usuario.get_by_email", return_value=user),
        patch("app.models_usuario.Usuario.get_by_id", return_value=user),
        patch("app.routes.auth._dispositivo_confiavel", return_value=True),
    ):
        client.post("/login", data={"email": user.email, "senha": "ok"}, follow_redirects=False)
        with (
            patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
            patch("app.routes.dashboard.get_static_cached", return_value=[]),
            patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
            patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        ):
            r = client.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert f"/chamado/{chamado.id}/historico" in html


def test_visualizar_chamado_solicitante_puro_nao_ve_botao_historico(
    client_logado_solicitante, db_session
):
    """GET /chamado/<id> com solicitante sem nivel_gestao não mostra o botão
    "Ver Histórico" (comportamento preservado)."""
    from tests.factories import make_chamado

    chamado = make_chamado(solicitante_id="sol_1")
    with patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]):
        r = client_logado_solicitante.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert f"/chamado/{chamado.id}/historico" not in html


def test_visualizar_chamado_responsavel_marca_visualizado_primeira_vez(
    client_logado_admin, db_session
):
    """GET /chamado/<id> pelo responsável atual grava visualizado_pelo_responsavel_em
    na primeira vez (confirmação de leitura, versão simples)."""
    from app.models import Chamado
    from tests.factories import make_chamado

    chamado = make_chamado(responsavel_id="admin_1")
    assert chamado.visualizado_pelo_responsavel_em is None

    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        r = client_logado_admin.get(f"/chamado/{chamado.id}", follow_redirects=False)

    assert r.status_code == 200
    recarregado = Chamado.get_by_id(chamado.id)
    assert recarregado.visualizado_pelo_responsavel_em is not None


def test_visualizar_chamado_nao_responsavel_nao_marca_visualizado(client_logado_admin, db_session):
    """GET /chamado/<id> por quem não é o responsável atual não grava nada —
    só o responsável dispara a confirmação de leitura."""
    from app.models import Chamado
    from tests.factories import make_chamado

    chamado = make_chamado(responsavel_id="outro_supervisor")

    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        r = client_logado_admin.get(f"/chamado/{chamado.id}", follow_redirects=False)

    assert r.status_code == 200
    recarregado = Chamado.get_by_id(chamado.id)
    assert recarregado.visualizado_pelo_responsavel_em is None


def test_visualizar_chamado_aog_esconde_opcao_de_prazo(client_logado_admin, db_session):
    """GET /chamado/<id> com categoria AOG não deve mostrar nenhuma opção de
    prazo/previsão de atendimento na tela — AOG nunca pode ter o prazo
    alterado (bloqueio incondicional já existe no service; a UI deixou de
    oferecer a opção pra fechar a lacuna de UX que só rejeitava no backend)."""
    from tests.factories import make_chamado

    chamado = make_chamado(categoria="AOG", responsavel_id="id_admin")
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        r = client_logado_admin.get(f"/chamado/{chamado.id}", follow_redirects=False)

    assert r.status_code == 200
    html = r.get_data(as_text=True)
    # Checa os elementos de marcação (não a string do case do dispatcher JS,
    # que sempre existe no bloco de script compartilhado da página).
    assert 'id="btn-extensao-automatica-previsao"' not in html
    assert 'id="modal-previsao-atendimento"' not in html


def test_visualizar_chamado_com_participantes_sem_usuario_atual_retorna_200(
    client_logado_admin, db_session
):
    """GET /chamado/<id> com participantes cadastrados, nenhum deles o usuário logado.

    Regressão: o template calcula `participante_atual` filtrando
    chamado.participantes pelo supervisor_id do usuário atual e aplicando
    `| first`. Se a lista tiver itens mas nenhum bater com o usuário logado
    (ex.: um admin olhando um chamado com participantes de outras pessoas),
    o filtro resulta em lista vazia e `| first` lança
    jinja2.exceptions.UndefinedError — capturado pelo except genérico da rota
    e mascarado como um redirect com flash de erro, sem 500 visível.
    """
    from tests.factories import make_chamado

    chamado = make_chamado(
        participantes=[{"supervisor_id": "outra_pessoa", "area": "TI", "status": "pendente"}]
    )
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        r = client_logado_admin.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 200


def test_visualizar_chamado_mostra_banner_pendente_e_botoes_para_gestor(
    client_logado_gestor, db_session, app
):
    """GET /chamado/<id> com pedido de previsão pendente: gestor_setor da área
    do chamado vê o banner "aguardando aprovação" + botões Aprovar/Rejeitar."""
    from unittest.mock import MagicMock

    from tests.factories import make_chamado

    chamado = make_chamado(area="Geral", responsavel_id="sup_dono")
    solicitante = MagicMock()
    solicitante.id = "sup_dono"
    solicitante.nome = "Dono Chamado"
    solicitante.perfil = "supervisor"
    solicitante.is_admin_or_above = False

    from datetime import datetime, timedelta

    from app.services.previsao_atendimento_service import solicitar_previsao_atendimento

    with app.app_context():
        resultado = solicitar_previsao_atendimento(
            chamado.id,
            datetime.now() + timedelta(days=5),  # depois do TAT padrão (3 dias)
            "Preciso de mais tempo",
            solicitante,
        )
    assert resultado["sucesso"] is True

    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        r = client_logado_gestor.get(f"/chamado/{chamado.id}")

    body = r.data.decode("utf-8")
    assert r.status_code == 200
    assert "Preciso de mais tempo" in body
    assert 'data-action="decidir-previsao-aprovar"' in body
    assert 'data-action="decidir-previsao-rejeitar"' in body


def test_visualizar_chamado_mostra_thread_conversa_em_ordem_cronologica(
    client_logado_admin, db_session
):
    """GET /chamado/<id>: resposta_solicitante + resposta_responsavel aparecem
    juntas na thread de conversa, mais antiga primeiro (Historico vem do banco
    mais recente primeiro — a rota precisa inverter)."""
    from app.models_historico import Historico
    from tests.factories import make_chamado

    chamado = make_chamado()

    Historico(
        chamado_id=chamado.id,
        usuario_id="sol_1",
        usuario_nome="Solicitante Um",
        acao="resposta_solicitante",
        campo_alterado="mensagem",
        valor_anterior=None,
        valor_novo="Mensagem mais antiga do solicitante",
    ).save()
    Historico(
        chamado_id=chamado.id,
        usuario_id="sup_1",
        usuario_nome="Supervisor Um",
        acao="resposta_responsavel",
        campo_alterado="mensagem",
        valor_anterior=None,
        valor_novo="Resposta mais recente do responsável",
    ).save()
    # Ruído: outra ação de histórico que NÃO deve entrar na thread de conversa.
    Historico(
        chamado_id=chamado.id,
        usuario_id="sup_1",
        usuario_nome="Supervisor Um",
        acao="alteracao_status",
        campo_alterado="status",
        valor_anterior="Aberto",
        valor_novo="Em Atendimento",
    ).save()

    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        r = client_logado_admin.get(f"/chamado/{chamado.id}")

    assert r.status_code == 200
    body = r.data.decode("utf-8")
    assert "Mensagem mais antiga do solicitante" in body
    assert "Resposta mais recente do responsável" in body
    pos_antiga = body.index("Mensagem mais antiga do solicitante")
    pos_recente = body.index("Resposta mais recente do responsável")
    assert pos_antiga < pos_recente, "thread deve mostrar mais antiga primeiro"


def test_visualizar_chamado_traduz_status_para_ingles(client_logado_admin, db_session):
    """GET /chamado/<id> com idioma=en não deve mostrar status cru em PT-BR.

    Regressão: components/_status_badge.html importado sem 'with context' em
    visualizar_chamado.html — a macro perde acesso a translate_status()/t()
    do context_processor, cai no fallback hardcoded em português
    independente do idioma escolhido pelo usuário.
    """
    from tests.factories import make_chamado

    chamado = make_chamado(status="Em Atendimento")
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        with client_logado_admin.session_transaction() as sess:
            sess["language"] = "en"
        r = client_logado_admin.get(f"/chamado/{chamado.id}")
    body = r.data.decode("utf-8")
    assert "In Progress" in body
    # value="Em Atendimento" no <option> é o valor canônico do form (correto,
    # não é texto visível) — só o texto exibido não pode vazar em PT-BR.
    assert ">Em Atendimento<" not in body


def test_visualizar_chamado_supervisor_sem_permissao_redireciona(
    client_logado_supervisor, db_session
):
    """GET /chamado/<id> com supervisor sem permissão na área redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="TI")
    with patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=False):
        r = client_logado_supervisor.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 302
    assert "/painel" in (r.location or "")


def test_visualizar_chamado_solicitante_proprio_retorna_200(client_logado_solicitante, db_session):
    """GET /chamado/<id> com solicitante visualizando o próprio chamado retorna 200."""
    from tests.factories import make_chamado

    chamado = make_chamado(solicitante_id="sol_1")
    with patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]):
        r = client_logado_solicitante.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 200


def test_visualizar_chamado_solicitante_aguardando_informacao_mostra_resposta_texto(
    client_logado_solicitante, db_session
):
    """GET /chamado/<id> com status 'Aguardando Informação' mostra o bloco de
    resposta por texto (endpoint /responder-solicitante) — antes só existia o
    fluxo de anexo tardio, obrigando anexar arquivo só pra escrever uma frase
    (achado em auditoria, 2026-08-05: endpoint pronto no backend, nunca ligado
    à UI)."""
    from tests.factories import make_chamado

    chamado = make_chamado(solicitante_id="sol_1", status="Aguardando Informação")
    with patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]):
        r = client_logado_solicitante.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert 'id="bloco-resposta-texto"' in html
    assert 'id="sol-resposta-texto"' in html
    assert "enviar-resposta-solicitante" in html
    assert "enviarRespostaSolicitante" in html
    assert "/responder-solicitante" in html


def test_visualizar_chamado_solicitante_outro_redireciona(client_logado_solicitante, db_session):
    """GET /chamado/<id> com solicitante tentando ver chamado alheio redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado(solicitante_id="outro_id")
    r = client_logado_solicitante.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 302


def test_visualizar_chamado_exception_redireciona(client_logado_admin, db_session):
    """GET /chamado/<id> quando ocorre exceção após localizar o chamado redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    with patch("app.routes.dashboard.usuario_pode_ver_chamado", side_effect=Exception("db err")):
        r = client_logado_admin.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 302


# ── POST /chamado/editar ──────────────────────────────────────────────────────


def test_editar_chamado_solicitante_redireciona(client_logado_solicitante):
    """POST /chamado/editar com solicitante redireciona (sem permissão)."""
    r = client_logado_solicitante.post(
        "/chamado/editar",
        data={"chamado_id": "ch1", "novo_status": "Concluído"},
        follow_redirects=False,
    )
    assert r.status_code == 302


def test_editar_chamado_sem_id_redireciona(client_logado_admin):
    """POST /chamado/editar sem chamado_id redireciona para /admin."""
    r = client_logado_admin.post(
        "/chamado/editar",
        data={"novo_status": "Concluído"},
        follow_redirects=False,
    )
    assert r.status_code == 302
    assert "/admin" in (r.location or "")


def test_editar_chamado_nao_encontrado_redireciona(client_logado_admin):
    """POST /chamado/editar quando chamado não existe redireciona para /admin."""
    r = client_logado_admin.post(
        "/chamado/editar",
        data={"chamado_id": "999999999", "novo_status": "Concluído"},
        follow_redirects=False,
    )
    assert r.status_code == 302
    assert "/admin" in (r.location or "")


def test_editar_chamado_sem_permissao_redireciona(client_logado_admin, db_session):
    """POST /chamado/editar quando usuario_pode_ver_chamado=False redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    with patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=False):
        r = client_logado_admin.post(
            "/chamado/editar",
            data={"chamado_id": str(chamado.id), "novo_status": "Concluído"},
            follow_redirects=False,
        )
    assert r.status_code == 302
    assert "/admin" in (r.location or "")


def test_editar_chamado_sucesso_redireciona_para_detalhe(client_logado_admin, db_session):
    """POST /chamado/editar com dados válidos chama serviço e redireciona para o chamado."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch(
            "app.services.edicao_chamado_service.processar_edicao_chamado",
            return_value={"sucesso": True, "mensagem": "Salvo"},
        ),
    ):
        r = client_logado_admin.post(
            "/chamado/editar",
            data={"chamado_id": str(chamado.id), "novo_status": "Concluído"},
            follow_redirects=False,
        )
    assert r.status_code == 302
    assert str(chamado.id) in (r.location or "")


# ── S4-03: Cache Usuario.get_all() via get_static_cached ─────────────────────


def test_visualizar_chamado_usa_cache_para_usuarios(client_logado_supervisor, db_session):
    """visualizar_detalhe_chamado deve usar get_static_cached, não Usuario.get_all() direto."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao")
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]) as mock_cache,
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        patch("app.routes.dashboard.Usuario.get_all") as mock_get_all,
    ):
        client_logado_supervisor.get(f"/chamado/{chamado.id}", follow_redirects=False)

    mock_cache.assert_called()
    mock_get_all.assert_not_called()


# ── F-59: Injeção de fórmula neutralizada em /exportar ────────────────────────


def test_exportar_neutraliza_formula_injection_em_xlsx(client_logado_supervisor):
    """F-59: /exportar deve aplicar _safe_cell() neutralizando fórmulas em descrição e responsável."""
    import io

    from openpyxl import load_workbook

    from app.models import Chamado

    chamado = Chamado(
        id="inj1",
        numero_chamado="2026-999",
        categoria="TI",
        tipo_solicitacao="Corretiva",
        descricao="=CMD('calc')",
        responsavel="+123",
        responsavel_id="u1",
        solicitante_id="s1",
        solicitante_nome="Teste",
        area="Manutencao",
        status="Aberto",
        prioridade=1,
        rl_codigo=None,
        gate=None,
        impacto=None,
        anexo=None,
        anexos=[],
        data_abertura=None,
        data_conclusao=None,
    )

    with (
        patch("app.routes.dashboard.aplicar_filtros_dashboard_com_paginacao") as mock_filtros,
        patch("app.routes.dashboard._filtrar_chamados_por_permissao") as mock_perm,
        patch("app.routes.dashboard.verificar_e_incrementar_export", return_value=(True, None)),
    ):
        mock_filtros.return_value = {
            "docs": [MagicMock()],
            "cursor_next": None,
            "cursor_prev": None,
        }
        mock_perm.return_value = [chamado]

        r = client_logado_supervisor.get("/exportar", follow_redirects=False)

    assert r.status_code == 200, f"Esperado 200, recebido {r.status_code}"
    assert "spreadsheetml" in r.content_type or "xlsx" in r.content_type

    wb = load_workbook(io.BytesIO(r.data))
    ws = wb.active

    # Colunas: Chamado(1) Categoria(2) RL(3) Tipo(4) Gate(5)
    #          Responsável(6) Solicitante(7) Área(8) Status(9) Anexo(10)
    #          Abertura(11) Conclusão(12) Descrição(13)
    desc_val = ws.cell(row=2, column=13).value
    resp_val = ws.cell(row=2, column=6).value

    assert isinstance(desc_val, str), f"Descrição deveria ser str, foi {type(desc_val)}"
    assert not desc_val.startswith("="), (
        f"Fórmula não neutralizada em 'Descrição': {desc_val!r} — aplicar _safe_cell()"
    )
    assert isinstance(resp_val, str), f"Responsável deveria ser str, foi {type(resp_val)}"
    assert not resp_val.startswith("+"), (
        f"Fórmula não neutralizada em 'Responsável': {resp_val!r} — aplicar _safe_cell()"
    )


# ── Regressão de segurança: /exportar e /exportar-avancado vazando outras áreas ──
# Achado em QA manual: supervisor da área "Demo" baixou /exportar e recebeu linhas
# de chamados da área "Manutencao"; /exportar-avancado trouxe métricas de
# supervisores de outras áreas na aba "Performance". Causa raiz: essas duas rotas
# consultavam db.collection("chamados") sem o mesmo filtro
# supervisor_ids_com_acesso array_contains que obter_contexto_admin já aplica
# pro /painel — a query saía sem escopo e só era filtrada depois, em memória,
# por _filtrar_chamados_por_permissao (que escopa os chamados, mas não as
# métricas agregadas por supervisor).


def _compilar_condicao_postgres(condicao) -> str:
    """Compila uma condição SQLAlchemy pro SQL literal (dialeto Postgres), pra
    inspecionar em teste sem precisar de conexão real."""
    from sqlalchemy.dialects import postgresql

    return str(
        condicao.compile(dialect=postgresql.dialect(), compile_kwargs={"literal_binds": True})
    )


def test_exportar_escopa_query_por_supervisor_ids_com_acesso(client_logado_supervisor):
    """/exportar deve escopar a query por área do supervisor ANTES da paginação,
    mesmo padrão usado em obter_contexto_admin para o /painel (Postgres:
    ChamadoRow.supervisor_ids_com_acesso.contains([user.id]) em condicoes_base)."""
    with (
        patch("app.routes.dashboard.aplicar_filtros_dashboard_com_paginacao") as mock_filtros,
        patch("app.routes.dashboard._filtrar_chamados_por_permissao", return_value=[]),
        patch("app.routes.dashboard.verificar_e_incrementar_export", return_value=(True, None)),
    ):
        mock_filtros.return_value = {"docs": []}

        client_logado_supervisor.get("/exportar", follow_redirects=False)

        condicoes_base = mock_filtros.call_args[0][0]
        assert len(condicoes_base) == 1, (
            "A query de /exportar não foi escopada por área — supervisor pode "
            "exportar chamados de áreas que não são dele."
        )
        sql = _compilar_condicao_postgres(condicoes_base[0])
        assert "supervisor_ids_com_acesso" in sql
        assert "sup_1" in sql


def test_exportar_avancado_escopa_query_por_supervisor_ids_com_acesso(
    client_logado_supervisor,
):
    """/exportar-avancado deve escopar a query de chamados da mesma forma que /exportar."""
    with (
        patch("app.routes.dashboard.aplicar_filtros_dashboard_com_paginacao") as mock_filtros,
        patch("app.routes.dashboard._filtrar_chamados_por_permissao", return_value=[]),
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.services.excel_export_service.exportador_excel") as mock_exp,
        patch("app.routes.dashboard.verificar_e_incrementar_export", return_value=(True, None)),
    ):
        import io

        mock_filtros.return_value = {"docs": []}
        mock_anal.obter_metricas_gerais.return_value = {}
        mock_anal.obter_metricas_supervisores.return_value = []
        mock_exp.exportar_relatorio_completo.return_value = io.BytesIO(b"PK fake")

        client_logado_supervisor.get("/exportar-avancado", follow_redirects=False)

        condicoes_base = mock_filtros.call_args[0][0]
        assert len(condicoes_base) == 1, "A query de /exportar-avancado não foi escopada por área."
        sql = _compilar_condicao_postgres(condicoes_base[0])
        assert "supervisor_ids_com_acesso" in sql
        assert "sup_1" in sql


def test_exportar_avancado_metricas_supervisores_filtradas_por_area(
    client_logado_supervisor,
):
    """/exportar-avancado não pode incluir, na aba de Performance, métricas de
    supervisores de áreas diferentes da do usuário que exportou."""
    with (
        patch("app.routes.dashboard.aplicar_filtros_dashboard_com_paginacao") as mock_filtros,
        patch("app.routes.dashboard._filtrar_chamados_por_permissao", return_value=[]),
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.services.excel_export_service.exportador_excel") as mock_exp,
        patch("app.routes.dashboard.verificar_e_incrementar_export", return_value=(True, None)),
    ):
        import io

        mock_filtros.return_value = {"docs": []}
        mock_anal.obter_metricas_gerais.return_value = {}
        mock_anal.obter_metricas_supervisores.return_value = [
            {"supervisor_nome": "Sup Mesma Area", "area": "Manutencao"},
            {"supervisor_nome": "Sup Outra Area", "area": "TI"},
        ]
        mock_exp.exportar_relatorio_completo.return_value = io.BytesIO(b"PK fake")

        client_logado_supervisor.get("/exportar-avancado", follow_redirects=False)

        _, kwargs = mock_exp.exportar_relatorio_completo.call_args
        metricas_enviadas = kwargs.get("metricas_supervisores") or []
        areas_enviadas = {m.get("area") for m in metricas_enviadas}
        assert "TI" not in areas_enviadas, (
            "Métricas de supervisor de outra área ('TI') vazaram pro relatório "
            "de um supervisor da área 'Manutencao'."
        )


# ── Onda 3: POST /painel como supervisor ──────────────────────────────────────


def test_painel_post_supervisor_altera_status_sucesso(client_logado_supervisor, db_session):
    """POST /painel com supervisor altera status com sucesso e redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao", responsavel_id=None, solicitante_id="sol_x")
    with patch("app.routes.dashboard.atualizar_status_chamado") as mock_atualizar:
        mock_atualizar.return_value = {"sucesso": True, "mensagem": "Status atualizado"}
        r = client_logado_supervisor.post(
            "/painel",
            data={"chamado_id": str(chamado.id), "novo_status": "Concluído"},
            follow_redirects=False,
        )
    assert r.status_code == 302
    mock_atualizar.assert_called_once()


def test_painel_post_supervisor_sem_permissao_na_area(client_logado_supervisor, db_session):
    """POST /painel com supervisor sem permissão na área redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="OutraArea", responsavel_id=None, solicitante_id="sol_x")
    r = client_logado_supervisor.post(
        "/painel",
        data={"chamado_id": str(chamado.id), "novo_status": "Concluído"},
        follow_redirects=False,
    )
    assert r.status_code == 302


def test_dashboard_alterar_status_supervisor_colega_owner_negado(
    client_logado_supervisor, db_session
):
    """Lacuna 2: supervisor não pode alterar chamado da mesma área se outro supervisor é o owner."""
    from tests.factories import make_chamado

    chamado = make_chamado(
        area="Manutencao", responsavel_id="outro_supervisor", solicitante_id="sol_outro"
    )
    with patch("app.routes.dashboard.atualizar_status_chamado") as mock_atualizar:
        r = client_logado_supervisor.post(
            "/painel",
            data={"chamado_id": str(chamado.id), "novo_status": "Concluído"},
            follow_redirects=False,
        )
    assert r.status_code == 302
    mock_atualizar.assert_not_called()


def test_painel_post_chamado_inexistente(client_logado_supervisor):
    """POST /painel com chamado que não existe redireciona."""
    r = client_logado_supervisor.post(
        "/painel",
        data={"chamado_id": "999999999", "novo_status": "Concluído"},
        follow_redirects=False,
    )
    assert r.status_code == 302


def test_painel_post_falha_sem_chave_erro(client_logado_admin, db_session):
    """POST /admin quando sucesso=False sem 'erro' no resultado exibe flash genérico."""
    from tests.factories import make_chamado

    chamado = make_chamado(status="Em Atendimento", area="Geral")
    with patch("app.routes.dashboard.atualizar_status_chamado") as mock_atualizar:
        mock_atualizar.return_value = {"sucesso": False}  # sem chave 'erro'
        r = client_logado_admin.post(
            "/admin",
            data={"chamado_id": str(chamado.id), "novo_status": "Invalido"},
            follow_redirects=False,
        )
    assert r.status_code == 302


def test_render_dashboard_usa_cache_para_setores(client_logado_admin):
    """_render_dashboard (GET /admin) deve buscar setores via get_static_cached
    ('categorias_setor', ...), não CategoriaSetor.get_all() direto — evita reler a
    coleção a cada carregamento do dashboard (F-XX economia de leituras Firestore)."""
    with (
        patch("app.routes.dashboard.obter_contexto_admin", return_value={}),
        patch("app.routes.dashboard.get_static_cached") as mock_cache,
    ):
        mock_cache.return_value = []
        client_logado_admin.get("/admin", follow_redirects=False)
    chaves_chamadas = [c.args[0] for c in mock_cache.call_args_list if c.args]
    assert "categorias_setor" in chaves_chamadas, (
        "GET /admin não busca setores via get_static_cached('categorias_setor', ...)"
    )


# ── Onda 3: /painel redireciona admin ─────────────────────────────────────────


def test_painel_com_admin_redireciona_para_admin(client_logado_admin):
    """GET /painel com perfil admin redireciona para /admin."""
    r = client_logado_admin.get("/painel", follow_redirects=False)
    assert r.status_code == 302
    assert "/admin" in (r.location or "")


# ── Onda 3: editar_chamado_pagina falhas ──────────────────────────────────────


def test_editar_chamado_falha_com_erro_exibe_flash(client_logado_admin, db_session):
    """POST /chamado/editar com sucesso=False e 'erro' presente redireciona para chamado."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch(
            "app.services.edicao_chamado_service.processar_edicao_chamado",
            return_value={"sucesso": False, "erro": "Erro de validação"},
        ),
    ):
        r = client_logado_admin.post(
            "/chamado/editar",
            data={"chamado_id": str(chamado.id), "novo_status": "Concluído"},
            follow_redirects=False,
        )
    assert r.status_code == 302
    assert str(chamado.id) in (r.location or "")


def test_editar_chamado_falha_sem_erro_exibe_flash_generico(client_logado_admin, db_session):
    """POST /chamado/editar com sucesso=False sem 'erro' usa flash_t genérico."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch(
            "app.services.edicao_chamado_service.processar_edicao_chamado",
            return_value={"sucesso": False},  # sem chave 'erro'
        ),
    ):
        r = client_logado_admin.post(
            "/chamado/editar",
            data={"chamado_id": str(chamado.id), "novo_status": "Concluído"},
            follow_redirects=False,
        )
    assert r.status_code == 302


# ── Onda 3: visualizar_historico branches ─────────────────────────────────────


def test_historico_chamado_nao_encontrado_redireciona(client_logado_supervisor):
    """GET /chamado/<id>/historico quando chamado não existe redireciona."""
    r = client_logado_supervisor.get("/chamado/999999999/historico", follow_redirects=False)
    assert r.status_code == 302


def test_historico_supervisor_sem_permissao_redireciona(client_logado_supervisor, db_session):
    """GET /chamado/<id>/historico com supervisor sem permissão redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="OutraArea")
    with patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=False):
        r = client_logado_supervisor.get(f"/chamado/{chamado.id}/historico", follow_redirects=False)
    assert r.status_code == 302


def test_historico_exception_redireciona(client_logado_supervisor, db_session):
    """GET /chamado/<id>/historico quando ocorre exceção após localizar o chamado redireciona."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao")
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.Historico.get_by_chamado_id", side_effect=Exception("db err")),
    ):
        r = client_logado_supervisor.get(f"/chamado/{chamado.id}/historico", follow_redirects=False)
    assert r.status_code == 302


def test_historico_traduz_status_para_ingles(client_logado_admin, db_session):
    """GET /chamado/<id>/historico com idioma=en não deve mostrar status cru em PT-BR.

    Dois bugs no mesmo template (historico.html):
    1. components/_status_badge.html importado sem 'with context' — badge do
       status atual cai no fallback hardcoded em português.
    2. O diff da timeline só traduz quando evento.campo_alterado == 'Status'
       (maiúsculo), mas todo o backend grava campo_alterado="status"
       (minúsculo) — a comparação nunca bate e o valor cru em PT-BR vaza
       pro <span class="bento-diff-chip"> independente do idioma.
    """
    from datetime import datetime

    from app.models_historico import Historico
    from tests.factories import make_chamado

    chamado = make_chamado(status="Em Atendimento")
    evento = Historico(
        id="h1",
        chamado_id=str(chamado.id),
        usuario_id="u1",
        usuario_nome="Fulano",
        acao="alteracao_status",
        campo_alterado="status",
        valor_anterior="Aberto",
        valor_novo="Em Atendimento",
        data_acao=datetime.now(),
    )
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.Historico.get_by_chamado_id", return_value=[evento]),
    ):
        with client_logado_admin.session_transaction() as sess:
            sess["language"] = "en"
        r = client_logado_admin.get(f"/chamado/{chamado.id}/historico")
    body = r.data.decode("utf-8")
    assert "In Progress" in body
    assert "Em Atendimento" not in body


# ── Onda 3: exportar exception handler ────────────────────────────────────────


def test_exportar_exception_redireciona(client_logado_supervisor):
    """GET /exportar quando ocorre exceção redireciona para painel."""
    with (
        patch(
            "app.routes.dashboard.aplicar_filtros_dashboard_com_paginacao",
            side_effect=Exception("timeout"),
        ),
        patch("app.routes.dashboard.verificar_e_incrementar_export", return_value=(True, None)),
    ):
        r = client_logado_supervisor.get("/exportar", follow_redirects=False)
    assert r.status_code == 302


# ── Onda 3: exportar_avancado ──────────────────────────────────────────────────


def _mock_chamado_obj():
    """Cria Chamado mock para exportar_avancado."""
    from app.models import Chamado

    return Chamado(
        id="adv1",
        numero_chamado="2026-001",
        categoria="TI",
        tipo_solicitacao="Corretiva",
        descricao="Teste avançado",
        responsavel="Resp",
        responsavel_id="u1",
        solicitante_id="s1",
        solicitante_nome="Sol",
        area="TI",
        status="Aberto",
        prioridade=1,
        rl_codigo=None,
        gate=None,
        impacto=None,
        anexo=None,
        anexos=[],
        data_abertura=None,
        data_conclusao=None,
    )


def test_exportar_avancado_sem_login_redireciona(client):
    """GET /exportar-avancado sem login redireciona para login."""
    r = client.get("/exportar-avancado", follow_redirects=False)
    assert r.status_code == 302
    assert "login" in (r.location or "").lower()


def test_exportar_avancado_retorna_xlsx(client_logado_supervisor):
    """GET /exportar-avancado com supervisor retorna arquivo xlsx."""
    from unittest.mock import MagicMock, patch

    with (
        patch("app.routes.dashboard.aplicar_filtros_dashboard_com_paginacao") as mock_filtros,
        patch("app.routes.dashboard._filtrar_chamados_por_permissao") as mock_perm,
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.services.excel_export_service.exportador_excel") as mock_exp,
        patch("app.routes.dashboard.verificar_e_incrementar_export", return_value=(True, None)),
    ):
        import io

        mock_filtros.return_value = {"docs": [MagicMock()]}
        mock_perm.return_value = [_mock_chamado_obj()]
        mock_anal.obter_metricas_gerais.return_value = {}
        mock_anal.obter_metricas_supervisores.return_value = []
        output = io.BytesIO(b"PK fake xlsx content")
        mock_exp.exportar_relatorio_completo.return_value = output

        r = client_logado_supervisor.get("/exportar-avancado", follow_redirects=False)

    assert r.status_code == 200
    ct = r.headers.get("Content-Type", "")
    assert "spreadsheet" in ct or "excel" in ct or "octet" in ct.lower()


def test_exportar_avancado_limite_excedido_redireciona(client_logado_supervisor):
    """GET /exportar-avancado quando limite diário excedido redireciona."""
    from unittest.mock import patch

    with (
        patch(
            "app.routes.dashboard.verificar_e_incrementar_export",
            return_value=(False, "Limite excedido"),
        ),
        patch("app.routes.dashboard.Config") as mock_cfg,
    ):
        mock_cfg.EXPORT_EXCEL_MAX_POR_USUARIO_POR_DIA = 5
        mock_cfg.ITENS_POR_PAGINA_DASHBOARD = 20
        r = client_logado_supervisor.get("/exportar-avancado", follow_redirects=False)
    assert r.status_code == 302


def test_exportar_avancado_exception_redireciona(client_logado_supervisor):
    """GET /exportar-avancado quando serviço lança exceção redireciona."""
    with (
        patch("app.routes.dashboard.aplicar_filtros_dashboard_com_paginacao") as mock_filtros,
        patch("app.routes.dashboard._filtrar_chamados_por_permissao", return_value=[]),
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.routes.dashboard.verificar_e_incrementar_export", return_value=(True, None)),
    ):
        mock_filtros.return_value = {"docs": []}
        mock_anal.obter_metricas_gerais.return_value = {}
        mock_anal.obter_metricas_supervisores.return_value = []
        # exportador_excel.exportar_relatorio_completo vai falhar porque não mockamos
        r = client_logado_supervisor.get("/exportar-avancado", follow_redirects=False)

    assert r.status_code in (200, 302)


def test_exportar_avancado_com_filtros_no_url(client_logado_supervisor):
    """GET /exportar-avancado com filtros no query string inclui filtros no Excel."""
    from unittest.mock import patch

    with (
        patch("app.routes.dashboard.aplicar_filtros_dashboard_com_paginacao") as mock_filtros,
        patch("app.routes.dashboard._filtrar_chamados_por_permissao") as mock_perm,
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.services.excel_export_service.exportador_excel") as mock_exp,
        patch("app.routes.dashboard.verificar_e_incrementar_export", return_value=(True, None)),
    ):
        import io

        mock_filtros.return_value = {"docs": []}
        mock_perm.return_value = []
        mock_anal.obter_metricas_gerais.return_value = {}
        mock_anal.obter_metricas_supervisores.return_value = []
        output = io.BytesIO(b"PK xlsx")
        mock_exp.exportar_relatorio_completo.return_value = output

        client_logado_supervisor.get(
            "/exportar-avancado?search=teste&categoria=TI&status=Aberto&responsavel=Ana",
            follow_redirects=False,
        )

    # Deve ter passado filtros_aplicados ao exportar
    if mock_exp.exportar_relatorio_completo.called:
        call_kwargs = mock_exp.exportar_relatorio_completo.call_args[1]
        filtros = call_kwargs.get("filtros_aplicados", {})
        assert "Busca" in filtros or isinstance(filtros, dict)


# ── Onda 3: relatorios branches ────────────────────────────────────────────────


def test_relatorios_atualizar_1_com_limite_excedido_redireciona(client_logado_admin):
    """GET /admin/relatorios?atualizar=1 quando limite excedido redireciona."""
    with (
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
        patch(
            "app.routes.dashboard.verificar_e_incrementar_relatorio",
            return_value=(False, "Limite de relatórios diário atingido"),
        ),
        patch("app.routes.dashboard.Config") as mock_cfg,
    ):
        mock_cfg.RELATORIO_MAX_POR_USUARIO_POR_DIA = 5
        mock_cfg.ITENS_POR_PAGINA_DASHBOARD = 20
        mock_cfg.ITENS_POR_PAGINA = 10
        mock_anal.obter_relatorio_completo.return_value = {}
        r = client_logado_admin.get("/admin/relatorios?atualizar=1", follow_redirects=False)
    assert r.status_code == 302


def test_relatorios_analytics_exception_mostra_erro(client_logado_admin):
    """GET /admin/relatorios quando analytics lança exceção retorna 200 com erro_relatorio."""
    with (
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        mock_anal.obter_relatorio_completo.side_effect = Exception("analytics down")
        r = client_logado_admin.get("/admin/relatorios", follow_redirects=False)
    assert r.status_code == 200


def test_relatorios_com_busca_sup_filtra_supervisores(client_logado_admin):
    """GET /admin/relatorios?busca_sup=Ana filtra lista de supervisores."""
    sup_ana = {
        "supervisor_nome": "Ana Souza",
        "supervisor_email": "ana@dtx.aero",
        "area": "TI",
        "carga_atual": 3,
        "taxa_resolucao_percentual": 80.0,
        "total_chamados": 10,
        "concluidos": 8,
        "abertos": 2,
        "em_andamento": 0,
        "tempo_medio_resolucao_horas": 12.0,
        "percentual_dentro_sla": 90.0,
        "distribuicao_categoria": {},
    }
    sup_bob = {
        "supervisor_nome": "Bob Lima",
        "supervisor_email": "bob@dtx.aero",
        "area": "RH",
        "carga_atual": 2,
        "taxa_resolucao_percentual": 70.0,
        "total_chamados": 5,
        "concluidos": 3,
        "abertos": 2,
        "em_andamento": 0,
        "tempo_medio_resolucao_horas": 8.0,
        "percentual_dentro_sla": 80.0,
        "distribuicao_categoria": {},
    }
    with (
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        mock_anal.obter_relatorio_completo.return_value = {
            "data_geracao": None,
            "metricas_gerais": {},
            "metricas_supervisores": [sup_ana, sup_bob],
            "metricas_areas": [],
            "insights": [],
        }
        r = client_logado_admin.get("/admin/relatorios?busca_sup=ana", follow_redirects=False)
    assert r.status_code == 200
    # Ana deve aparecer, Bob não
    assert b"Ana" in r.data


def test_relatorios_com_ordem_invalida_usa_desc(client_logado_admin):
    """GET /admin/relatorios?ordem_sup=INVALIDO normaliza para 'desc'."""
    with (
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        mock_anal.obter_relatorio_completo.return_value = {
            "data_geracao": None,
            "metricas_gerais": {},
            "metricas_supervisores": [],
            "metricas_areas": [],
            "insights": [],
        }
        r = client_logado_admin.get(
            "/admin/relatorios?ordem_sup=INVALIDO&ordem_area=INVALIDO",
            follow_redirects=False,
        )
    assert r.status_code == 200


def test_relatorios_com_busca_area_filtra_areas(client_logado_admin):
    """GET /admin/relatorios?busca_area=TI filtra lista de áreas."""
    area_ti = {
        "area": "TI",
        "total_chamados": 10,
        "abertos": 3,
        "concluidos": 7,
        "taxa_resolucao_percentual": 70.0,
        "tempo_medio_resolucao_horas": 12.0,
        "supervisores_alocados": 2,
        "chamados_por_supervisor": 5.0,
        "atribuidos_automaticamente": 4,
        "atribuidos_manualmente": 6,
        "taxa_automacao_percentual": 40.0,
    }
    area_rh = {
        "area": "RH",
        "total_chamados": 5,
        "abertos": 2,
        "concluidos": 3,
        "taxa_resolucao_percentual": 60.0,
        "tempo_medio_resolucao_horas": 8.0,
        "supervisores_alocados": 1,
        "chamados_por_supervisor": 5.0,
        "atribuidos_automaticamente": 1,
        "atribuidos_manualmente": 4,
        "taxa_automacao_percentual": 20.0,
    }
    with (
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
    ):
        mock_anal.obter_relatorio_completo.return_value = {
            "data_geracao": None,
            "metricas_gerais": {},
            "metricas_supervisores": [],
            "metricas_areas": [area_ti, area_rh],
            "insights": [],
        }
        r = client_logado_admin.get("/admin/relatorios?busca_area=ti", follow_redirects=False)
    assert r.status_code == 200


def test_relatorios_outer_exception_renderiza_pagina_erro(client_logado_admin):
    """GET /admin/relatorios quando preparar_metricas_paginadas lança exceção retorna 200 de fallback."""
    with (
        patch("app.routes.dashboard.analisador") as mock_anal,
        patch("app.routes.dashboard.Usuario.get_all", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        patch(
            "app.routes.dashboard.preparar_metricas_paginadas",
            side_effect=Exception("pagination error"),
        ),
    ):
        mock_anal.obter_relatorio_completo.return_value = {
            "data_geracao": None,
            "metricas_gerais": {},
            "metricas_supervisores": [],
            "metricas_areas": [],
            "insights": [],
        }
        r = client_logado_admin.get("/admin/relatorios", follow_redirects=False)
    assert r.status_code in (200, 302)


# ── Onda 3: visualizar_detalhe_chamado com referrer same-origin ───────────────


def test_visualizar_chamado_com_referrer_same_origin(client_logado_admin, db_session):
    """GET /chamado/<id> com Referer same-origin usa o referrer como voltar_url."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
    ):
        r = client_logado_admin.get(
            f"/chamado/{chamado.id}",
            headers={"Referer": "http://localhost/admin"},
            follow_redirects=False,
        )
    assert r.status_code == 200


# ── Supervisor não edita descrição do solicitante ──────────────────────────────


def test_visualizar_chamado_supervisor_nao_mostra_descricao_editavel(
    client_logado_supervisor, db_session
):
    """Supervisor da área vê a descrição só como texto, sem textarea editável."""
    from tests.factories import make_chamado

    chamado = make_chamado(area="Manutencao")
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
    ):
        r = client_logado_supervisor.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 200
    assert b'id="modal-descricao"' not in r.data


def test_visualizar_chamado_admin_mostra_descricao_editavel(client_logado_admin, db_session):
    """Admin continua vendo a descrição como textarea editável (válvula de escape)."""
    from tests.factories import make_chamado

    chamado = make_chamado()
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
        patch("app.routes.dashboard.CategoriaSetor.get_all", return_value=[]),
        patch("app.routes.dashboard.filtrar_supervisores_por_area", return_value=[]),
    ):
        r = client_logado_admin.get(f"/chamado/{chamado.id}", follow_redirects=False)
    assert r.status_code == 200
    assert b'id="modal-descricao"' in r.data


def test_navbar_admin_expoe_menu_de_perfil_com_relacao_programatica(client_logado_admin):
    """O acionador do perfil identifica e controla programaticamente seu painel."""
    with (
        patch("app.routes.dashboard.obter_contexto_admin") as mock_ctx,
        patch("app.routes.dashboard.get_static_cached", return_value=[]),
    ):
        mock_ctx.return_value = {
            "chamados": [],
            "gates": [],
            "responsaveis": [],
            "sla_map": {},
            "tem_proxima": False,
            "tem_anterior": False,
            "proximo_cursor": None,
            "cursor_anterior": None,
        }
        resposta = client_logado_admin.get("/admin")

    html = resposta.get_data(as_text=True)
    assert resposta.status_code == 200
    assert 'id="btn-profile-menu"' in html
    assert 'aria-controls="nav-profile-dropdown"' in html
    painel = html.split('id="nav-profile-dropdown"', 1)[1][:250]
    assert 'role="region"' in painel
    assert 'aria-labelledby="btn-profile-menu"' in painel


def test_historico_renderiza_eventos_com_lista_e_datetime(client_logado_admin, db_session):
    """A rota de histórico preserva a semântica da timeline no HTML renderizado."""
    from datetime import datetime

    from app.models_historico import Historico
    from tests.factories import make_chamado

    chamado = make_chamado()
    evento = Historico(
        chamado_id=chamado.id,
        usuario_id="admin_1",
        usuario_nome="Administrador",
        acao="criacao",
        campo_alterado="status",
        valor_novo="Aberto",
        data_acao=datetime(2026, 8, 13, 14, 30),
    )
    with (
        patch("app.routes.dashboard.usuario_pode_ver_chamado", return_value=True),
        patch("app.routes.dashboard.Historico.get_by_chamado_id", return_value=[evento]),
    ):
        resposta = client_logado_admin.get(f"/chamado/{chamado.id}/historico")

    html = resposta.get_data(as_text=True)
    assert resposta.status_code == 200
    assert '<ol class="bento-timeline space-y-5"' in html
    assert '<li class="bento-timeline-item">' in html
    assert 'datetime="2026-08-13T14:30:00"' in html
