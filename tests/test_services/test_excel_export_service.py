"""Testes do serviço de exportação Excel (exportador_excel.exportar_relatorio_completo)."""

import io

import pytest

from app.services.excel_export_service import MAX_EXPORT_CHAMADOS, _safe_cell, exportador_excel


def test_max_export_chamados_constante():
    """MAX_EXPORT_CHAMADOS está definido e é inteiro positivo."""
    assert MAX_EXPORT_CHAMADOS > 0
    assert isinstance(MAX_EXPORT_CHAMADOS, int)


def test_exportar_relatorio_completo_lista_vazia_retorna_bytes():
    """exportar_relatorio_completo com lista vazia de chamados retorna BytesIO (arquivo xlsx)."""
    output = exportador_excel.exportar_relatorio_completo(
        chamados=[],
        metricas_gerais={},
        metricas_supervisores=[],
        filtros_aplicados={},
    )
    assert isinstance(output, io.BytesIO)
    data = output.getvalue()
    assert len(data) > 0
    assert data[:2] == b"PK"


def test_exportar_relatorio_completo_com_chamado_mock_retorna_bytes():
    """exportar_relatorio_completo com um Chamado real retorna BytesIO sem exceção."""
    from app.models import Chamado

    chamado = Chamado(
        id="c1",
        numero_chamado="2026-001",
        categoria="Manutencao",
        tipo_solicitacao="Corretiva",
        descricao="Teste",
        responsavel="João",
        responsavel_id="u1",
        solicitante_id="s1",
        solicitante_nome="Maria",
        area="Manutencao",
        status="Aberto",
        prioridade=1,
        rl_codigo=None,
        gate=None,
        impacto=None,
        anexo=None,
        anexos=[],
        data_abertura=None,
        data_conclusao=None,
    )
    output = exportador_excel.exportar_relatorio_completo(
        chamados=[chamado],
        metricas_gerais={"total": 1, "abertos": 1},
        metricas_supervisores=[{"supervisor_nome": "João", "total_chamados": 1}],
        filtros_aplicados={},
    )
    assert isinstance(output, io.BytesIO)
    assert len(output.getvalue()) > 0


# ── Testes de segurança: formula injection ────────────────────────────────────


@pytest.mark.parametrize(
    "valor_entrada,esperado_prefixo",
    [
        ('=HYPERLINK("evil.com","click")', "'"),
        ("+1", "'"),
        ("-DROP TABLE", "'"),
        ("@SUM(A1:A10)", "'"),
        ("\tDATA", "'"),
        ("\rNEWLINE", "'"),
    ],
    ids=["igual", "mais", "menos", "arroba", "tab", "cr"],
)
def test_safe_cell_neutraliza_formula(valor_entrada, esperado_prefixo):
    """_safe_cell deve prefixar com ' strings que iniciam com char de fórmula."""
    resultado = _safe_cell(valor_entrada)
    assert isinstance(resultado, str)
    assert resultado.startswith(esperado_prefixo)
    assert valor_entrada in resultado


@pytest.mark.parametrize(
    "valor_seguro",
    ["texto normal", "2026-001", "Concluído", 42, None, 3.14, ""],
    ids=["texto", "numero_chamado", "status", "int", "none", "float", "vazio"],
)
def test_safe_cell_nao_altera_valores_seguros(valor_seguro):
    """_safe_cell não deve modificar valores que não iniciam com char de fórmula."""
    assert _safe_cell(valor_seguro) == valor_seguro


# ── F-59: Casos específicos de injection documentados ────────────────────────


@pytest.mark.parametrize(
    "valor_injecao",
    [
        "=CMD('calc')",
        "+123",
    ],
    ids=["cmd_injection", "mais_numerico"],
)
def test_safe_cell_neutraliza_casos_documentados_f59(valor_injecao):
    """F-59: _safe_cell neutraliza casos específicos de formula injection (=CMD, +123)."""
    resultado = _safe_cell(valor_injecao)
    assert resultado.startswith("'"), (
        f"'{valor_injecao}' deve ser prefixado com ', recebeu: {resultado!r}"
    )
    assert valor_injecao in resultado


# ── i18n: relatório Excel deve respeitar o idioma passado ────────────────────


def _mock_chamado(status="Concluído", categoria="Nao Aplicavel", tipo_solicitacao="TI"):
    from app.models import Chamado

    return Chamado(
        id="c1",
        numero_chamado="2026-001",
        categoria=categoria,
        tipo_solicitacao=tipo_solicitacao,
        descricao="Teste",
        responsavel="João",
        responsavel_id="u1",
        solicitante_id="s1",
        solicitante_nome="Maria",
        area="TI",
        status=status,
        prioridade=1,
        rl_codigo=None,
        gate=None,
        impacto=None,
        anexo=None,
        anexos=[],
        data_abertura=None,
        data_conclusao=None,
    )


def test_exportar_relatorio_completo_default_language_e_pt_br():
    """Sem passar language, o relatório continua em pt_BR (compatibilidade com chamadas antigas)."""
    from openpyxl import load_workbook

    output = exportador_excel.exportar_relatorio_completo(
        chamados=[_mock_chamado()],
        metricas_gerais={},
        metricas_supervisores=[],
        filtros_aplicados={},
    )
    wb = load_workbook(output)
    ws = wb.worksheets[1]  # aba "Chamados"
    cabecalho = [c.value for c in ws[1]]
    assert "Status" in cabecalho
    assert "Responsável" in cabecalho


def test_exportar_relatorio_completo_traduz_cabecalhos_para_ingles():
    """Com language='en', os cabeçalhos da aba de chamados saem em inglês."""
    from openpyxl import load_workbook

    output = exportador_excel.exportar_relatorio_completo(
        chamados=[_mock_chamado()],
        metricas_gerais={},
        metricas_supervisores=[],
        filtros_aplicados={},
        language="en",
    )
    wb = load_workbook(output)
    ws = wb.worksheets[1]  # aba "Chamados"
    cabecalho = [c.value for c in ws[1]]
    assert cabecalho == [
        "Ticket",
        "Category",
        "Type",
        "Status",
        "Responsible",
        "Requester",
        "Area",
        "Priority",
        "Opening Date",
        "Closing Date",
        "Impact",
    ]


def test_exportar_relatorio_completo_traduz_valores_status_e_categoria():
    """O status e a categoria do chamado (chaves internas do banco) saem traduzidos, não em PT cru."""
    from openpyxl import load_workbook

    output = exportador_excel.exportar_relatorio_completo(
        chamados=[_mock_chamado(status="Concluído", categoria="Nao Aplicavel")],
        metricas_gerais={},
        metricas_supervisores=[],
        filtros_aplicados={},
        language="en",
    )
    wb = load_workbook(output)
    ws = wb.worksheets[1]  # aba "Chamados"
    linha = [c.value for c in ws[2]]
    # índices: 0=numero, 1=categoria, 2=tipo, 3=status
    assert linha[1] == "Routine"
    assert linha[3] == "Completed"


def test_exportar_relatorio_completo_kpis_resumo_traduzidos_para_ingles():
    """Os rótulos de KPI na aba de resumo executivo saem traduzidos para o idioma pedido."""
    from openpyxl import load_workbook

    output = exportador_excel.exportar_relatorio_completo(
        chamados=[],
        metricas_gerais={"total_chamados": 5, "abertos": 2},
        metricas_supervisores=[],
        filtros_aplicados={},
        language="en",
    )
    wb = load_workbook(output)
    ws = wb.worksheets[0]  # aba "Resumo Executivo"
    valores_coluna_a = [c[0].value for c in ws.iter_rows(min_col=1, max_col=1) if c[0].value]
    assert "Total Tickets" in valores_coluna_a
    assert "Open" in valores_coluna_a
    assert "Abertos" not in valores_coluna_a


def test_exportar_relatorio_completo_status_sem_categoria_traduzido():
    """Chamado sem categoria usa o rótulo traduzido de 'sem categoria' na aba de análise."""
    from openpyxl import load_workbook

    output = exportador_excel.exportar_relatorio_completo(
        chamados=[_mock_chamado(categoria=None)],
        metricas_gerais={},
        metricas_supervisores=[],
        filtros_aplicados={},
        language="en",
    )
    wb = load_workbook(output)
    ws = wb.worksheets[4]  # aba "Categorias"
    valores_coluna_a = [c[0].value for c in ws.iter_rows(min_col=1, max_col=1) if c[0].value]
    assert "No Category" in valores_coluna_a
    assert "Sem Categoria" not in valores_coluna_a
