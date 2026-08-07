"""
Serviço Avançado de Exportação para Excel

Fornece exportação de chamados em formato XLSX com múltiplas abas,
formatação profissional, estilos, e análises.

**Funcionalidades:**

1. **Export Completo:**
   - Múltiplas abas: Resumo, Detalhes, Histórico
   - Formatação: cabeçalhos em cores, alinhamento, bordas
   - Styles: Cores indicando status/prioridade

2. **Configurações de Estilo:**
   - Cores padronizadas por tipo de dado (sucesso, alerta, info)
   - Fonte Calibri com tamanhos apropriados
   - Bordas e preenchimentos profissionais
   - Freeze panes para cabeçalhos

3. **Análise Histórica:**
   - Tempo de resolução (data abertura → conclusão)
   - Estatísticas por categoria/supervisor
   - Gráficos de tendência (pode ser adicionado com openpyxl)

**Uso Básico:**

```python
from app.services.excel_export_service import gerar_relatorio_excel

# Obter lista de chamados (usar MAX_EXPORT_CHAMADOS para limitar o tamanho do relatório)
chamados = db.collection('chamados').limit(MAX_EXPORT_CHAMADOS).stream()

# Gerar Excel
excel_bytes = gerar_relatorio_excel(
    chamados=list(chamados),
    tipo='completo',  # ou 'basico', 'analise'
    titulo='Relatório de Chamados - Fevereiro 2026'
)

# Enviar como download
response.data = excel_bytes
response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
response.headers['Content-Disposition'] = 'attachment; filename=chamados.xlsx'
return response
```

**Formatos Disponíveis:**
- `completo`: Todas as colunas e análises
- `basico`: Colunas essenciais apenas
- `analise`: Foco em histórico e duração
"""

import io
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.i18n import get_translated_category, get_translated_sector, get_translated_status
from app.i18n import get_translation as t

logger = logging.getLogger(__name__)

# Sentinela interna (não exibida) usada para agrupar chamados sem categoria —
# o rótulo traduzido (t('no_category_label')) só é aplicado na hora de escrever a célula.
_SEM_CATEGORIA = "Sem Categoria"

# Limite de chamados na exportação — mantém o relatório em tamanho razoável.
MAX_EXPORT_CHAMADOS = 100

# Chars que iniciam fórmulas em Excel/LibreOffice — prefixar com ' para neutralizar
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value: Any) -> Any:
    """Previne Excel formula injection em valores de células.
    Strings que iniciam com char de fórmula recebem aspas simples como prefixo,
    que o Excel interpreta como literal de texto (não executa a fórmula)."""
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


@dataclass
class ConfiguradorExcel:
    """Configuração de estilos e formatação para Excel"""

    # Cores
    COR_HEADER = "1F4E78"  # Azul escuro
    COR_TITULO = "2F5233"  # Verde escuro
    COR_ALERTA = "C65911"  # Laranja
    COR_SUCESSO = "70AD47"  # Verde
    COR_INFO = "4472C4"  # Azul

    # Fontes
    FONTE_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    FONTE_TITULO = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    FONTE_SUBTITULO = Font(name="Calibri", size=11, bold=True)
    FONTE_NORMAL = Font(name="Calibri", size=10)

    # Preenchimentos
    PREENCHIMENTO_HEADER = PatternFill(
        start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid"
    )
    PREENCHIMENTO_TITULO = PatternFill(
        start_color=COR_TITULO, end_color=COR_TITULO, fill_type="solid"
    )
    PREENCHIMENTO_LINHA_ALT = PatternFill(
        start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
    )

    # Bordas
    BORDA_PADRAO = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Alinhamentos
    ALINHAMENTO_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALINHAMENTO_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ALINHAMENTO_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


class ExportadorExcelAvancado:
    """Exportador profissional de relatórios em Excel"""

    def __init__(self):
        self.config = ConfiguradorExcel()

    def exportar_relatorio_completo(
        self,
        chamados: list[Any],
        metricas_gerais: dict[str, Any],
        metricas_supervisores: list[dict[str, Any]],
        filtros_aplicados: dict[str, str],
        language: str = "pt_BR",
    ) -> io.BytesIO:
        """Exporta relatório completo com múltiplas abas, no idioma informado."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove sheet padrão

        # Cria abas
        self._aba_resumo_executivo(wb, metricas_gerais, filtros_aplicados, language)
        self._aba_chamados_detalhados(wb, chamados, language)
        self._aba_performance_supervisores(wb, metricas_supervisores, language)
        self._aba_analise_status(wb, chamados, language)
        self._aba_analise_categorias(wb, chamados, language)

        # Salva em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _aba_resumo_executivo(
        self,
        wb: Workbook,
        metricas: dict[str, Any],
        filtros: dict[str, str],
        language: str,
    ) -> None:
        """Cria aba de resumo executivo com KPIs"""
        ws = wb.create_sheet(f"📊 {t('excel_sheet_summary', language)}", 0)
        ws.sheet_properties.tabColor = "1F4E78"

        # Configurar larguras das colunas
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25

        # Título
        ws["A1"] = t("excel_report_title", language)
        ws["A1"].font = self.config.FONTE_TITULO
        ws["A1"].fill = self.config.PREENCHIMENTO_TITULO
        ws["A1"].alignment = self.config.ALINHAMENTO_CENTER
        ws.merge_cells("A1:C1")
        ws.row_dimensions[1].height = 25

        # Data de geração
        agora = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws["A2"] = t("excel_generated_on", language, datetime=agora)
        ws["A2"].font = Font(name="Calibri", size=9, italic=True)
        ws.merge_cells("A2:C2")

        # Filtros aplicados
        if filtros:
            ws["A3"] = t("excel_filters_applied", language)
            ws["A3"].font = self.config.FONTE_SUBTITULO
            linha = 4
            for chave, valor in filtros.items():
                ws[f"A{linha}"] = f"  • {chave}: {valor}"
                ws[f"A{linha}"].font = self.config.FONTE_NORMAL
                linha += 1
            linha += 1
        else:
            linha = 4

        # KPIs Principais
        ws[f"A{linha}"] = t("excel_main_indicators", language)
        ws[f"A{linha}"].font = self.config.FONTE_SUBTITULO
        ws[f"A{linha}"].fill = self.config.PREENCHIMENTO_HEADER
        ws.merge_cells(f"A{linha}:C{linha}")

        linha += 1

        # Dados de KPI
        kpis = [
            (t("total_tickets", language), metricas.get("total_chamados", 0)),
            (t("open_tickets", language), metricas.get("abertos", 0)),
            (t("in_progress", language), metricas.get("em_andamento", 0)),
            (t("completed_tickets", language), metricas.get("concluidos", 0)),
            (
                t("resolution_rate", language),
                f"{metricas.get('taxa_resolucao_percentual', 0):.1f}%",
            ),
            (
                t("avg_resolution_time", language),
                f"{metricas.get('tempo_medio_resolucao_horas', 0):.1f}h",
            ),
        ]

        for chave, valor in kpis:
            ws[f"A{linha}"] = chave
            ws[f"A{linha}"].font = self.config.FONTE_NORMAL
            ws[f"B{linha}"] = valor
            ws[f"B{linha}"].font = Font(name="Calibri", size=10, bold=True)
            ws[f"B{linha}"].alignment = self.config.ALINHAMENTO_RIGHT

            # Adiciona borda
            for col in ["A", "B", "C"]:
                ws[f"{col}{linha}"].border = self.config.BORDA_PADRAO

            # Alternancia de cor
            if linha % 2 == 0:
                for col in ["A", "B", "C"]:
                    ws[f"{col}{linha}"].fill = self.config.PREENCHIMENTO_LINHA_ALT

            linha += 1

        # Distribuição por prioridade
        linha += 2
        ws[f"A{linha}"] = t("distribution_by_priority", language).upper()
        ws[f"A{linha}"].font = self.config.FONTE_SUBTITULO
        ws[f"A{linha}"].fill = self.config.PREENCHIMENTO_HEADER
        ws.merge_cells(f"A{linha}:C{linha}")

        linha += 1
        distribuicao = metricas.get("distribuicao_prioridade", {})
        for prioridade, quantidade in distribuicao.items():
            ws[f"A{linha}"] = t("excel_priority_n", language, priority=prioridade)
            ws[f"B{linha}"] = quantidade
            ws[f"B{linha}"].alignment = self.config.ALINHAMENTO_RIGHT

            for col in ["A", "B", "C"]:
                ws[f"{col}{linha}"].border = self.config.BORDA_PADRAO

            if linha % 2 == 0:
                for col in ["A", "B", "C"]:
                    ws[f"{col}{linha}"].fill = self.config.PREENCHIMENTO_LINHA_ALT

            linha += 1

    def _aba_chamados_detalhados(self, wb: Workbook, chamados: list[Any], language: str) -> None:
        """Cria aba com lista detalhada de chamados"""
        ws = wb.create_sheet(f"📋 {t('excel_sheet_tickets', language)}", 1)
        ws.sheet_properties.tabColor = "4472C4"

        # Colunas
        colunas = [
            (t("excel_col_ticket", language), 15),
            (t("category", language), 12),
            (t("type_label", language), 15),
            (t("status", language), 12),
            (t("responsible", language), 18),
            (t("requester", language), 15),
            (t("area_th", language), 12),
            (t("priority_label", language), 10),
            (t("opening_date_label", language), 15),
            (t("closing_date_label", language), 15),
            (t("impact_label", language), 10),
        ]

        # Header
        for col_num, (titulo, largura) in enumerate(colunas, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = largura

            cell = ws.cell(row=1, column=col_num, value=titulo)
            cell.font = self.config.FONTE_HEADER
            cell.fill = self.config.PREENCHIMENTO_HEADER
            cell.alignment = self.config.ALINHAMENTO_CENTER
            cell.border = self.config.BORDA_PADRAO

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # Dados
        for numero_linha, chamado in enumerate(chamados, 2):
            status_raw = chamado.status
            dados_linha = [
                chamado.numero_chamado,
                get_translated_category(chamado.categoria, language),
                get_translated_sector(chamado.tipo_solicitacao, language),
                get_translated_status(status_raw, language),
                chamado.responsavel,
                chamado.solicitante_nome or "-",
                chamado.area or "-",
                chamado.prioridade,
                chamado.data_abertura_formatada(),
                chamado.data_conclusao_formatada(),
                chamado.impacto or "-",
            ]

            for col_num, valor in enumerate(dados_linha, 1):
                cell = ws.cell(row=numero_linha, column=col_num, value=_safe_cell(valor))
                cell.font = self.config.FONTE_NORMAL
                cell.alignment = self.config.ALINHAMENTO_LEFT
                cell.border = self.config.BORDA_PADRAO

                # Alternancia de cor
                if numero_linha % 2 == 0:
                    cell.fill = self.config.PREENCHIMENTO_LINHA_ALT

                # Colorir status (compara o valor canônico do banco, não o rótulo traduzido)
                if col_num == 4:  # Status
                    if status_raw == "Concluído":
                        cell.font = Font(name="Calibri", size=10, color="70AD47", bold=True)
                    elif status_raw == "Aberto":
                        cell.font = Font(name="Calibri", size=10, color="C65911", bold=True)
                    elif status_raw == "Em Atendimento":
                        cell.font = Font(name="Calibri", size=10, color="4472C4", bold=True)

    def _aba_performance_supervisores(
        self, wb: Workbook, supervisores: list[dict[str, Any]], language: str
    ) -> None:
        """Cria aba de performance de supervisores"""
        ws = wb.create_sheet(f"👥 {t('excel_sheet_performance', language)}", 2)
        ws.sheet_properties.tabColor = "70AD47"

        # Colunas
        colunas = [
            (t("supervisor", language), 18),
            (t("total_assigned_label", language), 14),
            (t("completed_tickets", language), 12),
            (t("open_tickets", language), 12),
            (t("resolution_rate_pct_label", language), 16),
            (t("avg_time_hours_label", language), 14),
        ]

        for col_num, (titulo, largura) in enumerate(colunas, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = largura

            cell = ws.cell(row=1, column=col_num, value=titulo)
            cell.font = self.config.FONTE_HEADER
            cell.fill = self.config.PREENCHIMENTO_HEADER
            cell.alignment = self.config.ALINHAMENTO_CENTER
            cell.border = self.config.BORDA_PADRAO

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # Dados ordenados por taxa de resolução (decrescente)
        supervisores_ordenados = sorted(
            supervisores, key=lambda x: x.get("taxa_resolucao", 0), reverse=True
        )

        for numero_linha, sup in enumerate(supervisores_ordenados, 2):
            dados_linha = [
                sup.get("supervisor_nome", "N/A"),
                sup.get("total", 0),
                sup.get("concluidos", 0),
                sup.get("abertos", 0),
                round(sup.get("taxa_resolucao", 0), 1),
                round(sup.get("tempo_medio_resolucao", 0), 1),
            ]

            for col_num, valor in enumerate(dados_linha, 1):
                cell = ws.cell(row=numero_linha, column=col_num, value=_safe_cell(valor))
                cell.font = self.config.FONTE_NORMAL
                cell.border = self.config.BORDA_PADRAO

                # Alternancia de cor
                if numero_linha % 2 == 0:
                    cell.fill = self.config.PREENCHIMENTO_LINHA_ALT

                # Alinhar números à direita
                if col_num > 1:
                    cell.alignment = self.config.ALINHAMENTO_RIGHT
                else:
                    cell.alignment = self.config.ALINHAMENTO_LEFT

    def _aba_analise_status(self, wb: Workbook, chamados: list[Any], language: str) -> None:
        """Cria aba de análise por status"""
        ws = wb.create_sheet(f"📊 {t('status', language)}", 3)
        ws.sheet_properties.tabColor = "F79646"

        # Agrupar por status (chave canônica do banco — tradução só na hora de exibir)
        status_counts = {}

        for chamado in chamados:
            status_raw = chamado.status
            status_counts[status_raw] = status_counts.get(status_raw, 0) + 1

        # Header
        ws["A1"] = t("status_analysis_title", language).upper()
        ws["A1"].font = self.config.FONTE_TITULO
        ws["A1"].fill = self.config.PREENCHIMENTO_TITULO
        ws.merge_cells("A1:C1")

        ws["A2"] = t("status", language)
        ws["B2"] = t("quantity_label", language)
        ws["C2"] = t("percentage_label", language)

        for col in ["A", "B", "C"]:
            ws[f"{col}2"].font = self.config.FONTE_HEADER
            ws[f"{col}2"].fill = self.config.PREENCHIMENTO_HEADER
            ws[f"{col}2"].alignment = self.config.ALINHAMENTO_CENTER
            ws[f"{col}2"].border = self.config.BORDA_PADRAO

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12

        total = len(chamados)
        linha = 3

        for status_raw, quantidade in sorted(status_counts.items()):
            percentual = (quantidade / total * 100) if total > 0 else 0

            ws[f"A{linha}"] = get_translated_status(status_raw, language)
            ws[f"B{linha}"] = quantidade
            ws[f"C{linha}"] = f"{percentual:.1f}%"

            for col in ["A", "B", "C"]:
                ws[f"{col}{linha}"].border = self.config.BORDA_PADRAO
                ws[f"{col}{linha}"].font = self.config.FONTE_NORMAL

                if linha % 2 == 0:
                    ws[f"{col}{linha}"].fill = self.config.PREENCHIMENTO_LINHA_ALT

            # Alinhar números
            ws[f"B{linha}"].alignment = self.config.ALINHAMENTO_RIGHT
            ws[f"C{linha}"].alignment = self.config.ALINHAMENTO_RIGHT

            linha += 1

    def _aba_analise_categorias(self, wb: Workbook, chamados: list[Any], language: str) -> None:
        """Cria aba de análise por categoria"""
        ws = wb.create_sheet(f"🏷️ {t('nav_categories', language)}", 4)
        ws.sheet_properties.tabColor = "9966FF"

        # Agrupar por categoria (chave canônica do banco — tradução só na hora de exibir)
        categoria_counts = {}
        categoria_status = {}

        for chamado in chamados:
            cat = chamado.categoria or _SEM_CATEGORIA
            categoria_counts[cat] = categoria_counts.get(cat, 0) + 1

            if cat not in categoria_status:
                categoria_status[cat] = {}

            status_raw = chamado.status
            categoria_status[cat][status_raw] = categoria_status[cat].get(status_raw, 0) + 1

        # Header
        ws["A1"] = t("category_analysis_title", language).upper()
        ws["A1"].font = self.config.FONTE_TITULO
        ws["A1"].fill = self.config.PREENCHIMENTO_TITULO
        ws.merge_cells("A1:E1")

        ws["A2"] = t("category", language)
        ws["B2"] = t("total", language)
        ws["C2"] = t("open_tickets", language)
        ws["D2"] = t("in_progress", language)
        ws["E2"] = t("completed_tickets", language)

        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}2"].font = self.config.FONTE_HEADER
            ws[f"{col}2"].fill = self.config.PREENCHIMENTO_HEADER
            ws[f"{col}2"].alignment = self.config.ALINHAMENTO_CENTER
            ws[f"{col}2"].border = self.config.BORDA_PADRAO

        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 15

        linha = 3

        for categoria in sorted(categoria_counts.keys()):
            rotulo_categoria = (
                t("no_category_label", language)
                if categoria == _SEM_CATEGORIA
                else get_translated_category(categoria, language)
            )
            ws[f"A{linha}"] = rotulo_categoria
            ws[f"B{linha}"] = categoria_counts[categoria]
            ws[f"C{linha}"] = categoria_status[categoria].get("Aberto", 0)
            ws[f"D{linha}"] = categoria_status[categoria].get("Em Atendimento", 0)
            ws[f"E{linha}"] = categoria_status[categoria].get("Concluído", 0)

            for col in ["A", "B", "C", "D", "E"]:
                ws[f"{col}{linha}"].border = self.config.BORDA_PADRAO
                ws[f"{col}{linha}"].font = self.config.FONTE_NORMAL

                if linha % 2 == 0:
                    ws[f"{col}{linha}"].fill = self.config.PREENCHIMENTO_LINHA_ALT

                # Alinhar números
                if col != "A":
                    ws[f"{col}{linha}"].alignment = self.config.ALINHAMENTO_RIGHT

            linha += 1


# Instância global
exportador_excel = ExportadorExcelAvancado()
